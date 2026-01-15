import os
import sys
import json
import psutil

# 添加项目根目录到Python路径，以便正确导入模块
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

try:
    import pandas as pd
except ImportError:
    print("pandas未安装，请运行 pip install pandas 安装")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("openpyxl未安装，请运行 pip install openpyxl 安装")
    sys.exit(1)

from datetime import datetime
import calendar

# 动态导入数据库模块
def get_db_manager():
    from core.utils.database import get_db_manager
    return get_db_manager()


def ensure_result_folder():
    """确保项目根目录存在result文件夹"""
    # 获取脚本所在目录的父目录（项目根目录）
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    result_path = os.path.join(project_root, 'result')
    if not os.path.exists(result_path):
        os.makedirs(result_path)
    return result_path


def get_current_year_month():
    """获取当前年月，格式为yyyy-mm"""
    now = datetime.now()
    return now.strftime('%Y-%m')


def get_current_date():
    """获取当前日期，格式为yyyy-mm-dd"""
    now = datetime.now()
    return now.strftime('%Y-%m-%d')


def prepare_data_for_excel(collection_data):
    """
    准备数据以写入Excel
    :param collection_data: MongoDB中的数据
    :return: DataFrame格式的数据
    """
    if not collection_data or 'data' not in collection_data:
        # 如果没有数据，创建空的数据框架
        # 从12:00到23:00的12个小时
        hours = [f'{hour}:00' for hour in range(12, 24)]
        df = pd.DataFrame(columns=['门店名称'] + hours)
        return df

    # 提取所有门店名称
    all_shops = set()
    for hour_data in collection_data['data'].values():
        all_shops.update(hour_data.keys())
    
    all_shops = sorted(list(all_shops))  # 排序以保持一致的顺序
    
    # 创建数据框架
    hours = [f'{hour}:00' for hour in range(12, 24)]  # 12:00 到 23:00
    df_data = {'门店名称': all_shops}
    
    # 为每个小时填充数据
    for hour in hours:
        # 修复：处理数据库中小时格式的匹配
        hour_part = hour.split(':')[0]  # 提取小时部分，如 '0', '1', '2', '10', '11' 等
        # 数据库中存储的格式是两位数带前导零的格式
        hour_key = f"{int(hour_part):02d}"  # 转换为两位数格式，如 '0', '1' -> '00', '01'
        
        hour_column = []
        
        for shop in all_shops:
            # 检查该小时是否有此店铺的数据（尝试匹配数据库中的格式）
            if hour_key in collection_data['data'] and shop in collection_data['data'][hour_key]:
                hour_column.append(collection_data['data'][hour_key][shop])
            else:
                hour_column.append('0 / 0')  # 默认值
        
        df_data[hour] = hour_column
    
    df = pd.DataFrame(df_data)
    return df


def is_excel_running(file_path):
    """检查指定Excel文件是否正在运行"""
    try:
        # 首先检查文件是否存在
        if not os.path.exists(file_path):
            return False
        
        # 在Windows系统上，最可靠的方法是尝试打开文件进行写入
        # 如果无法打开，则文件很可能被Excel或其他程序占用
        import platform
        if platform.system().lower() == "windows":
            try:
                # 尝试以独占模式打开文件
                with open(file_path, 'r+b') as f:
                    pass  # 如果成功打开，说明文件未被占用
                return False
            except PermissionError:
                # 权限错误表示文件被占用
                return True
            except OSError:
                # 其他OSError也表明文件可能被占用
                return True
        else:
            # 对于非Windows系统，使用psutil检查
            try:
                for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                    try:
                        if proc.info['name'] and 'excel' in proc.info['name'].lower():
                            if proc.info['open_files']:
                                for open_file in proc.info['open_files']:
                                    # open_file可能是字符串或os.stat_result对象
                                    if hasattr(open_file, 'path'):
                                        file_path_str = open_file.path
                                    else:
                                        file_path_str = str(open_file)
                                        
                                    if file_path_str.lower() == file_path.lower():
                                        return True
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        continue
            except Exception:
                pass
            
            # 备用检测方法：尝试创建临时副本
            import tempfile
            import shutil
            try:
                temp_path = file_path + ".tmp_check"
                shutil.copy2(file_path, temp_path)
                os.remove(temp_path)
                return False
            except (PermissionError, OSError):
                return True
            
    except Exception as e:
        print(f"检查Excel进程时发生错误: {str(e)}")
        return False


def check_and_wait_for_excel(file_path, timeout=30):
    """检查Excel是否正在运行，如果是则等待或安全退出"""
    if not os.path.exists(file_path):
        return True  # 文件不存在，无需检查
    
    if not is_excel_running(file_path):
        return True  # Excel未运行，可以继续
    
    print(f"警告: Excel文件 {file_path} 正在运行中，请关闭后重试或等待自动重试...")
    
    # 等待一段时间让用户关闭Excel
    import time
    for i in range(timeout):
        time.sleep(1)
        if not is_excel_running(file_path):
            print("Excel文件已关闭，继续执行...")
            return True
        if i % 10 == 0 and i > 0:
            print(f"仍在等待Excel关闭... ({i}/{timeout}秒)")
    
    print(f"超时: Excel文件 {file_path} 仍然在运行，安全退出程序")
    return False


def sync_daily_data():
    """主函数：同步每日数据到Excel文件"""
    # 确保result文件夹存在
    result_folder = ensure_result_folder()
    
    # 获取当前年月作为文件名
    year_month = get_current_year_month()
    excel_filename = f"{year_month}.xlsx"
    excel_path = os.path.join(result_folder, excel_filename)
    
    # 检查Excel文件是否正在运行
    if not check_and_wait_for_excel(excel_path):
        print("由于Excel文件正在运行且超时，程序安全退出")
        return
    
    # 获取当前日期作为sheet名
    current_date = get_current_date()
    
    # 获取数据库管理器并连接
    db_manager = get_db_manager()
    if not db_manager.connect():
        print("无法连接到数据库")
        return
    
    try:
        # 从MongoDB获取当天的数据
        collection = db_manager.db['online_rate_new']
        date_data = collection.find_one({"sheet_date": current_date})
        
        # 准备要写入Excel的数据
        excel_data = prepare_data_for_excel(date_data)
        
        # 检查Excel文件是否存在，如果不存在则创建
        if not os.path.exists(excel_path):
            print(f"创建新的Excel文件: {excel_path}")
            # 创建一个新的Excel文件并将数据写入当前日期的sheet
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                excel_data.to_excel(writer, sheet_name=current_date, index=False)
            print(f"已创建文件: {excel_path} 并添加了 {current_date} 工作表，写入了 {len(excel_data)} 行数据")
        else:
            # 文件已存在，使用openpyxl直接操作
            workbook = load_workbook(excel_path)
            
            # 如果sheet已存在，删除它（避免重复）
            if current_date in workbook.sheetnames:
                del workbook[current_date]
            
            # 创建新的worksheet
            worksheet = workbook.create_sheet(title=current_date)
            
            # 将DataFrame转换为行并写入工作表
            for row in dataframe_to_rows(excel_data, index=False, header=True):
                worksheet.append(row)
            
            # 自动调整列宽和设置样式
            from openpyxl.styles import Alignment
            for idx, column in enumerate(worksheet.columns):
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass
                
                # 第一列保持原有宽度调整方式，其他列最小宽度设为100
                if idx == 0:  # 第一列
                    adjusted_width = min(max(max_length + 2, 10), 80)
                else:  # 其他列
                    adjusted_width = max(max_length + 2, 20)  # 最小宽度为30
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 为除第一列外的所有单元格设置居中对齐
                if idx > 0:
                    for cell in column:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置第一行（表头）的样式：黑色背景，白色字体，加粗
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')  # 黑色背景
            header_font = Font(color='FFFFFFFF', bold=True)  # 白色字体，加粗
            
            for cell in worksheet[1]:  # 获取第一行的所有单元格
                cell.fill = header_fill
                cell.font = header_font
                # 表头也设置居中对齐（包括第一列）
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 冻结第一行和第一列
            worksheet.freeze_panes = 'B2'
            
            # 保存工作簿
            workbook.save(excel_path)
            
            print(f"已更新 {excel_path} 中的 {current_date} 工作表，写入了 {len(excel_data)} 行数据")
    finally:
        # 断开数据库连接
        db_manager.disconnect()


if __name__ == "__main__":
    sync_daily_data()