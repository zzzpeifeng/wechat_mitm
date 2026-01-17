#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据Excel文件中的所有门店数据，生成一个按首次出现顺序排列的门店排序JSON文件
"""

import os
import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def extract_all_shops_from_excel(excel_path):
    """从Excel文件的所有sheet中提取门店名称"""
    workbook = load_workbook(excel_path)
    all_shops = set()
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # 假设门店名称在第一列（A列），从第二行开始是数据
        # 获取所有门店名称
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_idx == 0:  # 跳过标题行
                continue
            if row and len(row) > 0 and row[0]:  # 确保行不为空且第一列有值
                shop_name = str(row[0]).strip()
                if shop_name and shop_name != 'None':
                    all_shops.add(shop_name)
    
    return sorted(list(all_shops))  # 返回排序后的门店列表


def generate_shop_order_json(excel_path, output_path):
    """生成门店排序JSON文件"""
    shops = extract_all_shops_from_excel(excel_path)
    
    # 创建门店到索引的映射
    shop_order_dict = {}
    for idx, shop in enumerate(shops):
        shop_order_dict[shop] = idx
    
    # 保存为JSON文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(shop_order_dict, f, ensure_ascii=False, indent=2)
    
    print(f"已生成门店排序文件: {output_path}")
    print(f"总共包含 {len(shops)} 个门店")
    return shop_order_dict


def update_shop_order_from_multiple_excels(base_dir, output_path):
    """从多个Excel文件中收集所有门店并生成统一的排序JSON"""
    import glob
    
    all_shops = set()
    
    # 查找所有Excel文件
    excel_pattern = os.path.join(base_dir, "result", "*.xlsx")
    excel_files = glob.glob(excel_pattern)
    
    print(f"找到 {len(excel_files)} 个Excel文件:")
    for excel_file in excel_files:
        print(f"  - {os.path.basename(excel_file)}")
    
    # 从所有Excel文件中收集门店
    for excel_file in excel_files:
        try:
            workbook = load_workbook(excel_file)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 获取所有门店名称
                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if row_idx == 0:  # 跳过标题行
                        continue
                    if row and len(row) > 0 and row[0]:  # 确保行不为空且第一列有值
                        shop_name = str(row[0]).strip()
                        if shop_name and shop_name != 'None':
                            all_shops.add(shop_name)
        except Exception as e:
            print(f"处理文件 {excel_file} 时出错: {str(e)}")
            continue
    
    # 创建统一的门店排序
    shops = sorted(list(all_shops))  # 按字母顺序排序
    
    # 创建门店到索引的映射
    shop_order_dict = {}
    for idx, shop in enumerate(shops):
        shop_order_dict[shop] = idx
    
    # 保存为JSON文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(shop_order_dict, f, ensure_ascii=False, indent=2)
    
    print(f"已生成统一的门店排序文件: {output_path}")
    print(f"总共包含 {len(shops)} 个门店")
    return shop_order_dict


if __name__ == "__main__":
    # 获取项目根目录
    project_root = os.path.dirname(os.path.abspath(__file__))
    
    # Excel文件路径
    excel_path = os.path.join(project_root, 'result', '2026-01.xlsx')
    
    # 输出JSON文件路径
    output_path = os.path.join(project_root, 'shop_order.json')
    
    if os.path.exists(excel_path):
        print("使用单个Excel文件生成门店排序...")
        shop_order_dict = generate_shop_order_json(excel_path, output_path)
        print("门店排序字典:")
        for shop, idx in list(shop_order_dict.items())[:10]:  # 只打印前10个作为示例
            print(f"  {idx}: {shop}")
        if len(shop_order_dict) > 10:
            print(f"  ... 还有 {len(shop_order_dict) - 10} 个门店")
    else:
        print(f"Excel文件不存在: {excel_path}")
        
    print("\n" + "="*50)
    print("或者使用所有Excel文件生成统一的门店排序...")
    shop_order_dict = update_shop_order_from_multiple_excels(project_root, output_path)